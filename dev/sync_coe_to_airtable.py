#!/usr/bin/env python3
"""
Populate the Airtable Jobs table from the no-Actual-COE Salesforce pull.

    python3 dev/sync_coe_to_airtable.py --out "<folder>" [--dry-run] [--skip-report]

This replaces the Dynamics Export as the source of what the tracker contains. The
scope is the skill's, not this script's: run_report.py owns the SOQL, the
exclusions, the bucket/construction-state derivation, the duplicate reconciliation
and the verification pass. This reads the workbook that script produces and moves
it into Airtable. Two sources of truth for "which homesites are open work" would
drift apart, and the definitions are subtle enough (see the CofO note below) that
restating them here would be a second place to get them wrong.

Rules this script will not break:

  * It never deletes a row. A homesite that leaves the pull is marked
    Record Status = Closed with a Closed Date, and comes back to Active if it
    reappears. 66 of the rows leaving the current tracker carry hand-entered QA
    data -- walk dates, managers, key handover, notes -- that exists nowhere else
    in the company. Deleting them would be unrecoverable.

  * It only writes the fields in SF_OWNED. Everything a person typed is off
    limits, and assert_disjoint() fails the run rather than trusting the map.

  * It only writes fields that actually changed, so Last Synced keeps meaning
    "Salesforce data moved" rather than "the sync ran".

Requires the Salesforce CLI, so it runs on the Mac and not in a sandbox. Uses the
Python env run_report.py already builds (pandas + openpyxl).
"""

import argparse
import getpass
import glob
import json
import os
import socket
import subprocess
import sys
import time
import urllib.request
import urllib.error
from datetime import datetime, timezone

BASE_ID = 'appYX9df4lGO6G2uz'
JOBS_TABLE = 'tblqpmwtZ6i4gtogl'
SYNC_HISTORY_TABLE = 'tblBHVI7HelUb6vyk'
AIRTABLE_API = 'https://api.airtable.com/v0'

WORK = os.path.expanduser('~/.homesite_coe_report')
VENV_PY = os.path.join(WORK, 'venv', 'bin', 'python')

# The macOS system python3 has no openpyxl. run_report.py already builds a venv
# with pandas + openpyxl on first use, so borrow it rather than standing up a
# second environment that can drift from the one the report itself runs in.
if not os.environ.get('OLH_SYNC_REEXEC'):
    try:
        import openpyxl  # noqa: F401
    except ImportError:
        if not os.path.exists(VENV_PY):
            sys.exit('No openpyxl, and no venv at %s yet. Run the no-COE report once '
                     'first -- it builds the environment -- then re-run this.' % VENV_PY)
        os.environ['OLH_SYNC_REEXEC'] = '1'
        os.execv(VENV_PY, [VENV_PY, os.path.abspath(__file__)] + sys.argv[1:])


def find_skill_script():
    """Locate the skill's run_report.py.

    Its path contains session ids that change between runs, so it is discovered
    rather than hardcoded. --report-script overrides when that fails.
    """
    roots = [
        os.path.expanduser('~/Library/Application Support/Claude'),
        os.path.expanduser('~/.claude'),
    ]
    hits = []
    for root in roots:
        hits += glob.glob(os.path.join(
            root, '**', 'skills', 'homesites-no-actual-coe', 'scripts', 'run_report.py'),
            recursive=True)
    if not hits:
        return None
    # Newest wins: several sessions can leave copies behind.
    return max(hits, key=os.path.getmtime)

# ---------------------------------------------------------------------------
# Field map: workbook column -> Airtable field.
#
# Verified by measuring agreement against the 929 job numbers already in Airtable
# rather than by matching names, because two pairs are genuinely misleading:
#
#   Homesite Scheduled_Close_Date__c is LABELLED "Estimated COE Date" in
#   Salesforce. It is the estimate, not the actual. The skill warns that getting
#   this backwards inverts the whole report.
#
#   Airtable's "Scheduled Closing Date" is NOT that field and NOT an ECOE. It
#   comes from Opportunity.Scheduled_Closing_Date__c (91.7% agreement, versus
#   74.9% for the Homesite ECOE), and it disagrees with the ECOE on 129 of 784
#   rows -- often by weeks, not days. The tracker resolves urgency as
#   `Scheduled Closing Date || Estimated COE Date`, so feeding it an ECOE would
#   silently move the close date it sorts and flags on.
#
#   Homesite__c.CCC_Date__c is NOT where CCC gets recorded for OLH -- checked
#   2026-08-05 against live data (org sf-prod-observability-observer-claude):
#   0 of ~2,097 homesites in the OLH no-Actual-COE scope have it set, while
#   Opportunity.CCC_Date__c (same label, same type, verified to exist via
#   `sf sobject describe -s Opportunity`) is set on 431 of the 893 that have a
#   linked Opportunity at all. So CCC Date moved out of COLUMN_MAP (workbook,
#   Homesite-sourced, always null) and into EXTRA_FIELDS below, pulled via
#   Primary_Opportunity_ID__r the same way Scheduled Closing Date already is.
#   The remaining ~58% with no linked Opportunity (mostly unsold homesites)
#   will still show blank -- correctly, since there is no sale to hold a CCC
#   date yet.
# ---------------------------------------------------------------------------
COLUMN_MAP = {
    'Lot': 'Lot',
    'Homesite Street': 'Street Address',
    'City': 'City',
    'Zip': 'Zip',
    'Community': 'Community',
    'Homesite Status': 'Homesite Status',
    'Lot Status': 'Lot Status',
    'Bucket': 'Bucket',
    'Construction State': 'Construction State',
    'Sale Date': 'Sale Date',
    'Actual Start Date': 'Actual Start Date',
    'Constr Stage (JDE)': 'Construction Stage (JDE)',
    'Projected Completion': 'Projected Completion Date',
    'Actual Completion': 'Actual Completion Date',
    'Cert of Occupancy': 'Certificate of Occupancy Date',
    'Estimated COE Date': 'Estimated COE Date',
    'Construction Manager': 'Construction Manager',
    'Assigned Concierge': 'Assigned Concierge',
    'Salesforce Id': 'Salesforce Id',
    'Address Dup Check': 'Address Dup Check',
}

# Deliberately not synced from the workbook:
#   State                   constant 'FL' on all 1400 rows -- a column of noise.
#   Actual COE Date         null by definition; the pull is rows WITHOUT one.
#   JDE Sched Close (ECOE)  identical to Estimated COE in all 1378 rows where
#                           both are set, so it is a duplicate column.
#   Construction Stage      populated on 3 of 1400 rows in OLH.
#   CCC Date                always null on Homesite__c for OLH -- see above;
#                           now sourced from Opportunity via EXTRA_FIELDS.

DATE_COLUMNS = {
    'Sale Date', 'Actual Start Date', 'Projected Completion', 'Actual Completion',
    'Cert of Occupancy', 'Estimated COE Date',
}

# Fields the supplementary SOQL supplies, because the workbook does not carry them
# (or, for CCC Date, carries the wrong object's copy of them -- see the note above
# COLUMN_MAP).
#
# The kind is load bearing: a date-only comparison on a field that is really a
# dateTime would silently drop the time on every row it wrote. "Scheduled
# Closing Date" and "CCC Date" really are date-only.
#
# Construction Stage 7 (JDE) Date was dropped 2026-08-17 at Mauricio's request
# (removed from the sync and from Airtable) after a supplementary-pull failure
# investigation -- see git log for that date. The field itself still exists on
# Homesite__c and the query still works; it was removed as a product decision,
# not because the mapping went stale.
#
# Home Inspection Approved / Home Inspection Report Received / PHI Inspection
# Date added 2026-08-26 -- direct Homesite__c fields (no relationship hop
# needed, unlike Scheduled Closing Date / CCC Date above), added here rather
# than to the no-COE workbook's own FIELDS/COLUMN_MAP so this sync stays the
# single place new Salesforce columns get wired up without touching the
# separately-maintained homesites-no-actual-coe skill.
EXTRA_FIELDS = {
    'Scheduled Closing Date': ('Primary_Opportunity_ID__r.Scheduled_Closing_Date__c', 'date'),
    'CCC Date': ('Primary_Opportunity_ID__r.CCC_Date__c', 'date'),
    'Home Inspection Approved': ('HomeInspectionApproved__c', 'boolean'),
    'Home Inspection Report Received': ('HomeInspectionReportReceived__c', 'boolean'),
    'PHI Inspection Date': ('PHI_Inspection_Date__c', 'date'),
}

# ---------------------------------------------------------------------------
# Area Construction Manager: derived, not synced.
#
# The Completion Report filters on ACM, but Salesforce has no such field on
# Homesite__c -- the assignment is by community, and the mapping lives in
# dev/acm-map.json (built from the ACM.xlsx roster, 55 communities across 3
# ACMs). So it is computed here from the Community value this same sync writes,
# which keeps the two consistent by construction: a job that moves community
# gets the right ACM in the same pass.
#
# A community that is not in the map yields '' rather than a guess. Those show
# as blank on the report, which is the true state -- an unmapped community, not
# a homesite with no manager. Add it to acm-map.json when one appears.
# ---------------------------------------------------------------------------
ACM_FIELD = 'Area Construction Manager'
ACM_MAP_PATH = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'acm-map.json')


def load_acm_map():
    try:
        with open(ACM_MAP_PATH, encoding='utf-8') as fh:
            raw = json.load(fh)
    except FileNotFoundError:
        die('missing %s -- the Area Construction Manager mapping. Regenerate it '
            'from the ACM roster workbook rather than syncing the column blank.'
            % ACM_MAP_PATH)
    # Compare on a squashed key so trailing spaces or case in either source do
    # not silently drop a community.
    return {' '.join(str(k).split()).upper(): v for k, v in raw.items()}


ACM_BY_COMMUNITY = load_acm_map()


def acm_for(community):
    return ACM_BY_COMMUNITY.get(' '.join(str(community or '').split()).upper(), '')


# field -> 'date' | 'datetime' | 'text', for normalisation and comparison.
FIELD_KIND = {}
for _col, _field in COLUMN_MAP.items():
    FIELD_KIND[_field] = 'date' if _col in DATE_COLUMNS else 'text'
for _field, (_sf, _kind) in EXTRA_FIELDS.items():
    FIELD_KIND[_field] = _kind
FIELD_KIND[ACM_FIELD] = 'text'

SYNC_MANAGED = {'Record Status', 'Closed Date', 'Last Synced'}

SF_OWNED = set(COLUMN_MAP.values()) | set(EXTRA_FIELDS) | SYNC_MANAGED | {ACM_FIELD}

# Everything a person types. None of it comes from Salesforce.
MANUAL_FIELDS = {
    'QA Ready', 'QAI Date', 'QAI Manager', 'QAI Complete',
    'QAA Date', 'QAA Manager', 'QAA Accepted',
    'CEL Date', 'CEL Manager', 'CEL Completed', 'Buyer Attended CEL',
    'ACC Date', 'ACC Manager', 'ACC Completed', 'Buyer Attended ACC',
    'NOC Lock Date', 'Power Meter', 'Water Meter',
    'Construction Risk', 'Construction Risk Notes',
    'Land Risk', 'Land Risk Notes',
    'Key Status', 'Delivered To', 'Delivery Date', 'Notes', 'CEL Letter Sent',
}

# Exception to the manual/SF split above: Construction Risk and Land Risk stay
# hand-editable (a leader can check or clear either at any time), but the sync
# is allowed to check one -- never clear one -- when Salesforce says the risk
# is real and the box is not already checked. Airtable returns an unchecked
# checkbox as absent, the same as one that was never touched, so "blank" and
# "a leader cleared it" are indistinguishable; that ambiguity is exactly why
# this can only ever add a checkmark, never remove one. Sourced from
# Homesite__c.Construction_Risk__c / Land_Risk__c via fetch_risk_flags().
POPULATE_IF_BLANK = {'Construction Risk', 'Land Risk'}


def assert_disjoint():
    """A field cannot be both Salesforce-owned and hand-entered."""
    clash = SF_OWNED & MANUAL_FIELDS
    if clash:
        sys.exit('REFUSING TO RUN: these fields are listed as both '
                 'Salesforce-owned and manually maintained: %s' % ', '.join(sorted(clash)))


def die(msg):
    sys.exit('\nFAILED: %s\n' % msg)


# ---------------------------------------------------------------------------
# Airtable
# ---------------------------------------------------------------------------

def pat():
    v = os.environ.get('AIRTABLE_PAT', '').strip()
    if not v:
        die('AIRTABLE_PAT is not set. Try:\n'
            '  AIRTABLE_PAT=$(netlify env:get AIRTABLE_PAT | grep -o "pat[A-Za-z0-9._]*") \\\n'
            '    python3 dev/sync_coe_to_airtable.py --out "<folder>"')
    return v


def airtable(method, path, body=None, attempts=4):
    """One Airtable call, retried on transport failures and 429/5xx.

    Retries matter here because this is meant to run unattended every morning: a
    single dropped TLS handshake part-way through 1400 rows would otherwise leave
    the table half-written. A 4xx other than 429 is not retried -- that is a bug
    in the request, and repeating it just takes longer to tell us.
    """
    url = AIRTABLE_API + '/' + BASE_ID + path
    payload = json.dumps(body).encode() if body else None
    delay = 1.0
    for attempt in range(1, attempts + 1):
        req = urllib.request.Request(
            url, method=method, data=payload,
            headers={'Authorization': 'Bearer ' + pat(),
                     **({'Content-Type': 'application/json'} if body else {})})
        try:
            with urllib.request.urlopen(req, timeout=45) as r:
                raw = r.read().decode()
                return json.loads(raw) if raw else None
        except urllib.error.HTTPError as e:
            detail = e.read().decode()[:400]
            if e.code in (429, 500, 502, 503, 504) and attempt < attempts:
                print('    Airtable %d, retrying in %.0fs' % (e.code, delay), flush=True)
            else:
                die('Airtable %s on %s %s\n  %s' % (e.code, method, path, detail))
        except (urllib.error.URLError, TimeoutError, OSError) as e:
            if attempt >= attempts:
                die('could not reach Airtable after %d attempts on %s %s\n  %s'
                    % (attempts, method, path, e))
            print('    %s reaching Airtable, retrying in %.0fs'
                  % (type(e).__name__, delay), flush=True)
        time.sleep(delay)
        delay *= 2


def fetch_jobs():
    """Every row, keyed by Job #.

    Returns (by_job, blank, dupes). Rows with no Job # are reported separately so
    they are never silently archived -- a blank primary field is a data problem,
    not an absence. Duplicate Job # values are reported because Airtable does not
    enforce uniqueness on a primary field, so a duplicated table looks completely
    healthy until half the updates land on the copy nobody is looking at.
    """
    records, offset, blank, dupes = {}, None, [], {}
    while True:
        path = '/%s?pageSize=100%s' % (JOBS_TABLE, ('&offset=' + offset) if offset else '')
        page = airtable('GET', path)
        for r in page.get('records', []):
            job = str((r.get('fields') or {}).get('Job #', '')).strip()
            if not job:
                blank.append(r['id'])
            elif job in records:
                dupes.setdefault(job, [records[job]]).append(r)
            else:
                records[job] = r
        offset = page.get('offset')
        if not offset:
            break
    return records, blank, dupes


def write_batches(verb, payload, dry):
    """Airtable takes 10 records per request. PATCH for updates, POST for creates."""
    if not payload:
        return 0
    if dry:
        return len(payload)
    done = 0
    for i in range(0, len(payload), 10):
        chunk = payload[i:i + 10]
        airtable('PATCH' if verb == 'update' else 'POST', '/' + JOBS_TABLE,
                 {'records': chunk, 'typecast': True})
        done += len(chunk)
        if done % 100 == 0 or done == len(payload):
            print('    %s %d/%d' % (verb, done, len(payload)), flush=True)
        # Airtable allows 5 requests/second/base. A full run is ~145 requests, so
        # without this it spends most of its time being 429'd and retried.
        time.sleep(0.25)
    return done


def log_sync_history(div, started, finished, status, exit_code, reason, metrics):
    """Append one row to the Sync History table -- one row per run, success or
    not, so the tracker's sync-history page reflects what launchd actually did
    rather than a log file nobody but this machine can read.

    Deliberately never allowed to fail the run: if Airtable is unreachable
    here (the same failure the sync itself may have just hit), a missing log
    row is fine. Raising from inside logging and masking the sync's own real
    result would not be.
    """
    fields = {
        'Started': started.isoformat(),
        'Finished': finished.isoformat(),
        'Duration (sec)': round((finished - started).total_seconds()),
        'Status': status,
        'Exit Code': exit_code,
        'Reason': reason,
        'Division': div,
        'Origin': os.environ.get('OLH_SYNC_ORIGIN', 'manual'),
        'Host': socket.gethostname(),
        'Triggered By': getpass.getuser(),
    }
    for key, field in (('rows_raw', 'Rows Raw'), ('rows_final', 'Rows Final'),
                       ('airtable_total', 'Airtable Total Rows'),
                       ('airtable_active', 'Airtable Active Rows')):
        if metrics.get(key) is not None:
            fields[field] = metrics[key]

    try:
        token = os.environ.get('AIRTABLE_PAT', '').strip()
        if not token:
            print('WARNING: no AIRTABLE_PAT -- Sync History row not written.', flush=True)
            return
        req = urllib.request.Request(
            AIRTABLE_API + '/' + BASE_ID + '/' + SYNC_HISTORY_TABLE,
            method='POST',
            data=json.dumps({'records': [{'fields': fields}], 'typecast': True}).encode(),
            headers={'Authorization': 'Bearer ' + token, 'Content-Type': 'application/json'})
        with urllib.request.urlopen(req, timeout=20) as r:
            r.read()
    except Exception as e:
        print('WARNING: could not write Sync History row: %s' % e, flush=True)


# ---------------------------------------------------------------------------
# Salesforce
# ---------------------------------------------------------------------------

def run_report(out, division, alias, script=None):
    """Let the skill produce the workbook. Its verification pass is the gate: it
    exits non-zero when the data changed shape, and syncing an unverified pull is
    how a bad upstream day becomes 1400 wrong rows in the tracker."""
    script = script or find_skill_script()
    if not script or not os.path.exists(script):
        die('cannot find the no-COE skill\'s run_report.py.\n'
            'Pass --report-script <path>, or --skip-report to sync a workbook '
            'that is already in --out.')
    print('Running the no-COE report (it owns the scope and the definitions):', flush=True)
    print('  %s' % script, flush=True)
    p = subprocess.run([sys.executable, script, '--division', division,
                        '--out', out, '--alias', alias])
    if p.returncode != 0:
        die('run_report.py exited %d -- verification failed. Do not sync an '
            'unverified pull; read its FAIL lines first.' % p.returncode)


def supplementary(division, alias):
    """The fields the workbook does not carry, joined on Job #."""
    cols = [spec[0] for spec in EXTRA_FIELDS.values()]
    q = ('SELECT Name, %s FROM Homesite__c WHERE DivisionCode__c = \'%s\' '
         "AND Actual_COE_Date_New__c = NULL AND (Homesite_Status__c != 'Available' "
         'OR Actual_Start_Date__c != NULL OR Actual_Completion_Date__c != NULL)'
         % (', '.join(cols), division))
    print('Pulling %d supplementary field(s) from Salesforce...' % len(cols), flush=True)
    p = subprocess.run(['sf', 'data', 'query', '-o', alias, '-r', 'csv', '-q', q],
                       capture_output=True, text=True)
    if p.returncode != 0:
        die('sf data query failed:\n' + (p.stderr or '')[:600])
    import csv
    import io
    rows = list(csv.DictReader(io.StringIO(p.stdout)))
    out = {}
    for r in rows:
        job = (r.get('Name') or '').strip()
        if not job:
            continue
        rec = {}
        for air, (sf, kind) in EXTRA_FIELDS.items():
            raw = (r.get(sf) or '').strip()
            # Salesforce's CSV export renders booleans as the literal strings
            # "true"/"false" -- coerce to a real Python bool here, at the
            # source, so a checkbox field is never handed a string downstream
            # (the Airtable API wants JSON true/false, not "true").
            rec[air] = (raw.lower() == 'true') if kind == 'boolean' else raw
        out[job] = rec
    print('  %d rows' % len(out), flush=True)
    return out


def fetch_risk_flags(division, alias):
    """Construction_Risk__c / Land_Risk__c from Homesite__c, for POPULATE_IF_BLANK.

    Deliberately not folded into supplementary()/EXTRA_FIELDS: those feed SF_OWNED
    fields that get overwritten outright, and these two must never be. Kept as a
    separate query and a separate write path so a bug here can't silently widen
    into "overwrite" the way a shared code path might invite.
    """
    q = ("SELECT Name, Construction_Risk__c, Land_Risk__c FROM Homesite__c "
         "WHERE DivisionCode__c = '%s' AND Actual_COE_Date_New__c = NULL "
         "AND (Homesite_Status__c != 'Available' OR Actual_Start_Date__c != NULL "
         "OR Actual_Completion_Date__c != NULL)" % division)
    print('Pulling risk flags from Salesforce...', flush=True)
    p = subprocess.run(['sf', 'data', 'query', '-o', alias, '-r', 'csv', '-q', q],
                       capture_output=True, text=True)
    if p.returncode != 0:
        die('sf data query (risk flags) failed:\n' + (p.stderr or '')[:600])
    import csv
    import io
    rows = list(csv.DictReader(io.StringIO(p.stdout)))
    out = {}
    for r in rows:
        job = (r.get('Name') or '').strip()
        if not job:
            continue
        out[job] = {
            'Construction Risk': str(r.get('Construction_Risk__c') or '').strip().lower() == 'true',
            'Land Risk': str(r.get('Land_Risk__c') or '').strip().lower() == 'true',
        }
    print('  %d rows' % len(out), flush=True)
    return out


# ---------------------------------------------------------------------------
# Value handling
# ---------------------------------------------------------------------------

def norm_date(v):
    """Airtable wants YYYY-MM-DD. Blank means blank, never today."""
    if v is None or v == '':
        return ''
    if hasattr(v, 'strftime'):
        return v.strftime('%Y-%m-%d')
    s = str(v).strip()
    if not s:
        return ''
    for fmt in ('%Y-%m-%d', '%m/%d/%Y', '%Y-%m-%dT%H:%M:%S.%f%z', '%Y-%m-%dT%H:%M:%S%z'):
        try:
            return datetime.strptime(s[:26] if '%f' in fmt else s, fmt).strftime('%Y-%m-%d')
        except ValueError:
            continue
    return s[:10]


def norm_text(v):
    """Trim, and collapse internal whitespace runs to one space.

    Salesforce pads some text fields internally -- Construction Manager comes back
    as "Layton, Brian                      (OLH)" -- while the data already in
    Airtable is collapsed. Without this, 885 of 929 manager names "changed" on
    every run: cosmetically worse, endlessly churning Last Synced, and worse than
    cosmetic because the walk pages match roster members by exact name, so a
    padded copy silently stops matching.
    """
    if v is None:
        return ''
    return ' '.join(str(v).split())


def norm_bool(v):
    """Canonical True/False regardless of source shape.

    `have` (from Airtable) arrives as a real Python bool, or is absent entirely
    for an unchecked box (compare() already maps that absence to ''). `want`
    (from supplementary()) is already a real bool by the time it gets here too
    -- but this stays permissive of a stray "true"/"false" string so a future
    caller that skips that conversion fails closed (unrecognised -> False)
    rather than comparing a string to a bool and reporting a false change on
    every single run, the same class of churn norm_text() exists to prevent.
    """
    if isinstance(v, bool):
        return v
    return str(v).strip().lower() in ('true', '1', 'yes')


def read_workbook(path, division):
    import openpyxl
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    sheet = '%s No Actual COE' % division
    if sheet not in wb.sheetnames:
        die('sheet "%s" not in %s (found: %s)' % (sheet, path, ', '.join(wb.sheetnames)))
    ws = wb[sheet]
    rows = list(ws.iter_rows(values_only=True))
    header = [norm_text(h) for h in rows[0]]
    missing = [c for c in COLUMN_MAP if c not in header]
    if missing:
        die('the workbook is missing expected column(s): %s\n'
            'The export changed shape; update COLUMN_MAP rather than guessing.'
            % ', '.join(missing))
    idx = {h: i for i, h in enumerate(header)}
    out = {}
    for r in rows[1:]:
        job = norm_text(r[idx['Job #']])
        if not job:
            continue
        rec = {}
        for col, field in COLUMN_MAP.items():
            rec[field] = normalise(field, r[idx[col]])
        rec[ACM_FIELD] = acm_for(rec.get('Community'))
        out[job] = rec
    return out


def norm_datetime(v):
    """Airtable accepts ISO 8601; keep the time, normalise to UTC 'Z'."""
    if v is None or v == '':
        return ''
    if hasattr(v, 'isoformat'):
        return v.isoformat()
    s = str(v).strip()
    if not s:
        return ''
    s = s.replace('+0000', '+00:00')
    try:
        dt = datetime.fromisoformat(s)
    except ValueError:
        return s
    if dt.tzinfo is None:
        dt = dt.replace(tzinfo=timezone.utc)
    return dt.astimezone(timezone.utc).strftime('%Y-%m-%dT%H:%M:%S.000Z')


def normalise(field, v):
    kind = FIELD_KIND.get(field, 'text')
    if kind == 'date':
        return norm_date(v)
    if kind == 'datetime':
        return norm_datetime(v)
    if kind == 'boolean':
        return norm_bool(v)
    return norm_text(v)


def compare(desired, current):
    """Only the fields that actually differ.

    Both sides go through the same normaliser before comparison, so a field is
    reported as changed only when its VALUE moved -- not when Salesforce formats it
    differently from however it was last written. Getting this wrong is not
    harmless: it churns Last Synced (whose whole point is "Salesforce data moved"),
    and it rewrites 900 rows a day for nothing.

    Airtable omits empty fields rather than returning them, so absent and '' are
    the same thing here, and clearing a value means sending None.
    """
    delta = {}
    for field, want in desired.items():
        have = current.get(field, '')
        if have is None:
            have = ''
        if isinstance(have, list):          # linked records; never in SF_OWNED
            continue
        if normalise(field, have) != normalise(field, want):
            delta[field] = want if want != '' else None
    return delta


# ---------------------------------------------------------------------------
# main
# ---------------------------------------------------------------------------

def main():
    p = argparse.ArgumentParser()
    p.add_argument('--out', required=True, help='Folder the workbook is written to')
    p.add_argument('--division', default='OLH')
    p.add_argument('--alias', default='sf-prod-observability-observer-claude')
    p.add_argument('--dry-run', action='store_true',
                   help='Report what would change and write nothing')
    p.add_argument('--skip-report', action='store_true',
                   help='Use the workbook already in --out instead of re-pulling')
    p.add_argument('--report-script', default=None,
                   help="Path to the skill's run_report.py, if discovery fails")
    a = p.parse_args()

    assert_disjoint()
    div = a.division.upper()

    # Every run gets exactly one Sync History row, success or not -- see
    # log_sync_history(). status/exit_code/reason are set at whichever exit
    # point is actually hit; metrics is filled in as far as sync() gets before
    # any failure, so a row for a failed run can still carry a partial picture
    # (e.g. rows_raw/rows_final if the Salesforce pull succeeded but the
    # Airtable write did not).
    started = datetime.now(timezone.utc)
    metrics = {}
    status, exit_code, reason = 'Success', 0, ''

    # One run at a time, always.
    #
    # Two concurrent runs each read the table, each compute the same set of
    # creates, and each write them: 471 new homesites became 942 rows and 471
    # duplicate Job # values. Airtable does not enforce uniqueness on a primary
    # field, so nothing fails -- the table just quietly holds every homesite twice.
    # This happened for real, because a wrapper timed out and the run looked dead
    # while it was still writing.
    lock = None
    try:
        if not a.dry_run:
            os.makedirs(WORK, exist_ok=True)
            lock_path = os.path.join(WORK, 'sync-%s.lock' % div)
            try:
                lock = os.open(lock_path, os.O_CREAT | os.O_EXCL | os.O_WRONLY)
                os.write(lock, ('pid %d started %s\n'
                                % (os.getpid(), datetime.now().isoformat())).encode())
            except FileExistsError:
                try:
                    held = open(lock_path).read().strip()
                except OSError:
                    held = '(unreadable)'
                status, exit_code = 'Blocked', 1
                reason = 'Another sync for %s already running (lock collision): %s' % (div, held)
                die('another sync for %s is already running:\n  %s\n\n'
                    'Wait for it to finish. A slow run is normal -- ~1400 rows takes a '
                    'couple of minutes. If you are certain nothing is running, delete:\n'
                    '  %s' % (div, held, lock_path))

        sync(a, div, metrics)
    except SystemExit as e:
        # die() calls sys.exit(msg) with the message AS the exit code, not an
        # int -- that string is exactly the reason a person watching the log
        # would want. A plain int code (or None, for a clean exit that still
        # somehow lands here) falls back to a generic reason.
        if status == 'Success':  # not already set by the lock-collision branch above
            status = 'Failed'
        exit_code = 1
        reason = reason or (str(e.code) if e.code else 'Exited unexpectedly.')
    except Exception as e:
        status, exit_code = 'Failed', 1
        reason = reason or ('%s: %s' % (type(e).__name__, e))
    finally:
        if lock is not None:
            os.close(lock)
            try:
                os.unlink(os.path.join(WORK, 'sync-%s.lock' % div))
            except OSError:
                pass
        finished = datetime.now(timezone.utc)
        if not a.dry_run:
            log_sync_history(div, started, finished, status, exit_code, reason, metrics)

    if status != 'Success':
        sys.exit(exit_code)


def sync(a, div, metrics=None):
    if metrics is None:
        metrics = {}
    if not a.skip_report:
        run_report(a.out, div, a.alias, a.report_script)

    path = os.path.join(a.out, '%s Homesites - No Actual COE Date.xlsx' % div)
    if not os.path.exists(path):
        die('no workbook at %s' % path)

    print('\nReading %s' % path, flush=True)
    report = read_workbook(path, div)
    print('  %d homesites in scope' % len(report), flush=True)
    metrics['rows_final'] = len(report)

    extra = supplementary(div, a.alias)
    metrics['rows_raw'] = len(extra)
    for job, vals in extra.items():
        if job in report:
            report[job].update(vals)

    risk = fetch_risk_flags(div, a.alias)

    print('\nReading Airtable...', flush=True)
    current, blank, dupes = fetch_jobs()
    print('  %d rows (%d with no Job #, left alone)' % (len(current), len(blank)), flush=True)

    if dupes:
        die('%d Job # value(s) appear on more than one row, e.g. %s\n\n'
            'Syncing now would update one copy and leave the other stale, and the '
            'tracker would show both. Clean them up first:\n'
            '  node dev/dedupe-jobs.js --apply' % (len(dupes), ', '.join(list(dupes)[:5])))

    creates, updates, archives, reactivates = [], [], [], []
    now = datetime.now(timezone.utc).isoformat()
    today = datetime.now().strftime('%Y-%m-%d')

    risk_populated = {}

    for job, desired in report.items():
        rec = current.get(job)
        risk_flags = risk.get(job) or {}
        if rec is None:
            fields = {k: v for k, v in desired.items() if v not in ('', None)}
            fields['Job #'] = job
            fields['Record Status'] = 'Active'
            fields['Last Synced'] = now
            for f in POPULATE_IF_BLANK:
                if risk_flags.get(f):
                    fields[f] = True
                    risk_populated[f] = risk_populated.get(f, 0) + 1
            creates.append({'fields': fields})
            continue

        have = rec.get('fields') or {}

        # Manually archived by a human -- never touch again, even if the
        # job is still in (or comes back into) the Salesforce pull.
        if have.get('Manual Archive - Do Not Resync'):
            continue

        delta = compare(desired, have)

        # A homesite back in the pull is open work again.
        if have.get('Record Status') != 'Active':
            delta['Record Status'] = 'Active'
            delta['Closed Date'] = None
            reactivates.append(job)

        # Populate-if-blank: only ever add a checkmark, never remove one.
        for f in POPULATE_IF_BLANK:
            if risk_flags.get(f) and not have.get(f):
                delta[f] = True
                risk_populated[f] = risk_populated.get(f, 0) + 1

        if delta:
            delta['Last Synced'] = now
            updates.append({'id': rec['id'], 'fields': delta})

    for job, rec in current.items():
        if job in report:
            continue
        have = rec.get('fields') or {}
        if have.get('Record Status') == 'Closed':
            continue        # already archived; leave it alone
        archives.append({'id': rec['id'], 'fields': {
            'Record Status': 'Closed',
            'Closed Date': have.get('Closed Date') or today,
            'Last Synced': now,
        }})

    # Nothing hand-entered may appear in any payload, except the two
    # populate-if-blank risk flags, which are allowed by design (see
    # POPULATE_IF_BLANK) and are never overwritten -- only ever set True.
    for batch in (creates, updates, archives):
        for rec in batch:
            bad = (set(rec['fields']) & MANUAL_FIELDS) - POPULATE_IF_BLANK
            if bad:
                die('a write payload contains hand-entered field(s): %s' % ', '.join(sorted(bad)))
            for f in POPULATE_IF_BLANK:
                if rec['fields'].get(f) is False:
                    die('a write payload tries to clear %s -- populate-if-blank '
                        'may only ever set this True' % f)

    pad = lambda s, n: str(s).ljust(n)
    print('\n' + '=' * 62)
    print(pad('create (new to the tracker)', 44) + str(len(creates)))
    print(pad('update (Salesforce data moved)', 44) + str(len(updates)))
    print(pad('  of which reactivated', 44) + str(len(reactivates)))
    print(pad('archive (left the pull, row KEPT)', 44) + str(len(archives)))
    print(pad('unchanged', 44) + str(len(report) - len(creates) - len(updates)))
    print('=' * 62)

    if updates:
        counts = {}
        for rec in updates:
            for f in rec['fields']:
                if f == 'Last Synced':
                    continue
                counts[f] = counts.get(f, 0) + 1
        print('\nfields changing:')
        for f, n in sorted(counts.items(), key=lambda kv: -kv[1]):
            print('   ' + pad(f, 42) + str(n))

    if risk_populated:
        print('\nrisk flags populated (blank -> checked; never cleared):')
        for f, n in sorted(risk_populated.items(), key=lambda kv: -kv[1]):
            print('   ' + pad(f, 42) + str(n))

    if a.dry_run:
        print('\nDRY RUN -- nothing written.')
        return

    print('\nWriting...', flush=True)
    write_batches('update', archives, False)
    write_batches('update', updates, False)
    write_batches('create', creates, False)

    after, _, after_dupes = fetch_jobs()
    if after_dupes:
        print('\nWARNING: %d duplicate Job # value(s) exist after this run. Run '
              '`node dev/dedupe-jobs.js --apply`.' % len(after_dupes))
    active = sum(1 for r in after.values()
                 if (r.get('fields') or {}).get('Record Status') == 'Active')
    metrics['airtable_total'] = len(after)
    metrics['airtable_active'] = active
    print('\nDone. Airtable now holds %d rows, %d Active.' % (len(after), active))
    if active != len(report):
        print('NOTE: %d Active but %d in the pull. Every row in the pull is set Active '
              'and every row outside it is archived, so these should match. A gap means '
              'some writes did not land -- re-run and compare before trusting the '
              'tracker.' % (active, len(report)))


if __name__ == '__main__':
    main()
